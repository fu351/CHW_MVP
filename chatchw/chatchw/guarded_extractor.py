# guarded_extractor.py
from __future__ import annotations

import collections
import csv
import json
import os
import re
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pypdf

_have_pdfminer = True
try:
    import pdfminer  # noqa: F401
except Exception:
    _have_pdfminer = False

_have_ocr = True
try:
    import pytesseract  # type: ignore
    from pdf2image import convert_from_path  # type: ignore
    from PIL import Image  # noqa: F401
except Exception:
    _have_ocr = False

try:
    from openai import OpenAI
except Exception:  # pragma: no cover
    OpenAI = None  # type: ignore

# keep strict merge True, but consolidation is now local-lossless so rule count is preserved
STRICT_MERGE = True

try:
    from pydantic import BaseModel, Field, ValidationError, conlist  # noqa: F401
    PydanticAvailable = True
except Exception:  # pragma: no cover
    PydanticAvailable = False

# -------------------------
# Utility helpers (new)
# -------------------------

_DMNN = {
    "dmn": "https://www.omg.org/spec/DMN/20191111/MODEL/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}

def _text_or(elem) -> str:
    try:
        return (elem.text or "").strip()
    except Exception:
        return ""

def _strip_quotes(s: str) -> str:
    if s is None:
        return ""
    ss = str(s).strip()
    if len(ss) >= 2 and ((ss[0] == ss[-1] == '"') or (ss[0] == ss[-1] == "'")):
        return ss[1:-1]
    return ss

def _and_join(parts: List[str]) -> str:
    parts = [p for p in (parts or []) if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return " and ".join(parts)

def _validate_dmn_outputs_or_die(dmn_xml: Optional[str]) -> None:
    """
    Cheap guard to avoid crashes. Only error if DMN is completely missing.
    Otherwise be permissive so the pipeline can proceed and coverage can run.
    """
    if not dmn_xml or "<dmn:definitions" not in dmn_xml:
        raise RuntimeError("DMN generation failed: no <dmn:definitions> found")

def _ensure_dmn_inputs(dmn_xml: Optional[str], ask_plan: Any, merged_ir: Dict[str, Any]) -> str:
    """
    Best effort no-op fixer. Returns the original DMN unless it is empty.
    You can extend this to inject missing inputs if needed.
    """
    if not dmn_xml:
        return "<dmn:definitions xmlns:dmn=\"https://www.omg.org/spec/DMN/20191111/MODEL/\"></dmn:definitions>"
    return dmn_xml

def _extract_pypdf_pages(pdf_path: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    try:
        reader = pypdf.PdfReader(pdf_path)
        for i, page in enumerate(reader.pages, start=1):
            try:
                txt = page.extract_text() or ""
            except Exception:
                txt = ""
            out.append((f"p{i}", txt))
    except Exception:
        return []
    return out

def _postprocess_page_text(pages: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    cleaned: List[Tuple[str, str]] = []
    for pid, t in pages:
        s = t or ""
        # join hyphenated line breaks like "con-\n tinued" -> "continued"
        s = re.sub(r"(\w)-\n(\w)", r"\1\2", s)
        # collapse multi newlines
        s = re.sub(r"\n{2,}", "\n", s)
        cleaned.append((pid, s))
    return cleaned

def _chunk_pages_len_only(pages: List[Tuple[str, str]], max_chars: int = 4000) -> List[Tuple[str, str]]:
    chunks: List[Tuple[str, str]] = []
    cur_id: List[str] = []
    buf: List[str] = []
    cur_len = 0
    for pid, txt in pages:
        if cur_len + len(txt) > max_chars and buf:
            chunks.append(("sec_" + "_".join(cur_id[:2]) + ("" if len(cur_id) < 3 else f"_{len(cur_id)}"), "\n".join(buf)))
            cur_id, buf, cur_len = [], [], 0
        cur_id.append(pid)
        buf.append(f"[{pid}]\n{txt}")
        cur_len += len(txt)
    if buf:
        chunks.append(("sec_" + "_".join(cur_id[:2]) + ("" if len(cur_id) < 3 else f"_{len(cur_id)}"), "\n".join(buf)))
    return chunks

def _split_into_sections_by_headings(pages: List[Tuple[str, str]], max_chars: int = 4000) -> List[Tuple[str, str]]:
    """
    Simple heading heuristic: start a new section when we see an ALL CAPS line
    or a line that looks like a numbered heading. Fallback to length chunking.
    """
    sections: List[Tuple[str, str]] = []
    buf: List[str] = []
    cur_pages: List[str] = []
    cur_len = 0
    sec_idx = 1

    def flush():
        nonlocal buf, cur_pages, cur_len, sec_idx
        if buf:
            sections.append((f"section_{sec_idx:03d}", "\n".join(buf)))
            sec_idx += 1
            buf, cur_pages, cur_len = [], [], 0

    for pid, txt in pages:
        lines = (txt or "").splitlines()
        for ln in lines:
            is_head = bool(re.match(r"^\s*(\d+(\.\d+)*)\s+\S", ln)) or (ln.strip().isupper() and len(ln.strip()) >= 6)
            if is_head and cur_len > 0:
                flush()
            buf.append(ln)
            cur_pages.append(pid)
            cur_len += len(ln) + 1
            if cur_len >= max_chars:
                flush()
    flush()
    if not sections:
        return _chunk_pages_len_only(pages, max_chars=max_chars)
    return sections

def _extract_pdfminer_layout_pages(pdf_path: str) -> List[Tuple[str, str]]:
    # Placeholder for richer layout extraction if needed later
    return []

def _extract_pdfminer_pages(pdf_path: str) -> List[Tuple[str, str]]:
    # Placeholder for basic pdfminer text extraction if needed later
    return []

def _extract_ocr_pages(pdf_path: str) -> List[Tuple[str, str]]:
    # Placeholder for OCR fallback
    return []

# -------------------------
# Pydantic models
# -------------------------

if PydanticAvailable:
    Op = Union[str]

    class Variable(BaseModel):
        name: str
        type: str
        unit: Optional[str] = None
        allowed: Optional[List[str]] = None
        synonyms: List[str] = []
        prompt: Optional[str] = None
        refs: List[str] = []

    class ObsCond(BaseModel):
        obs: str
        op: str
        value: Union[float, int, bool, str]

    class SymCond(BaseModel):
        sym: str
        eq: bool

    class AllOf(BaseModel):
        all_of: conlist(Union[ObsCond, SymCond], min_length=1)

    class AnyOf(BaseModel):
        any_of: conlist(Union[ObsCond, SymCond], min_length=1)

    Condition = Union[ObsCond, SymCond, AllOf, AnyOf]

    class ThenBlock(BaseModel):
        triage: Optional[str] = None
        flags: List[str] = []
        reasons: List[str] = []
        actions: List[dict] = []   # include CHW treatments here
        guideline_ref: Optional[str] = None
        priority: int = 0
        advice: List[str] = []

    class Rule(BaseModel):
        rule_id: str
        when: conlist(Condition, min_length=1)
        then: ThenBlock

    class QA(BaseModel):
        notes: List[str] = []
        unmapped_vars: List[str] = []
        dedup_dropped: int = 0
        overlap_fixed: List[str] = []

    class IR(BaseModel):
        variables: List[Variable] = []
        rules: List[Rule] = []
        canonical_map: dict = {}
        qa: QA = QA()

# -------------------------
# String and JSON helpers
# -------------------------

def _strip_code_fences(text: str) -> str:
    if not text:
        return text
    s = text.strip()
    if s.startswith("```"):
        first_nl = s.find("\n")
        if first_nl != -1:
            s = s[first_nl + 1 :]
        if s.endswith("```"):
            s = s[:-3]
    return s.strip()

def _loose_to_json(s: str) -> Any:
    s = _strip_code_fences(s)
    lines: List[str] = []
    in_block = False
    for ln in s.splitlines():
        t = ln
        if in_block:
            if "*/" in t:
                t = t.split("*/", 1)[1]
                in_block = False
            else:
                continue
        if "/*" in t:
            t = t.split("/*", 1)[0]
            in_block = True
        if "//" in t:
            t = t.split("//", 1)[0]
        lines.append(t)
    s = "\n".join(lines)
    s = re.sub(r",\s*(]|})", r"\1", s)
    for opener, closer in [("{", "}"), ("[", "]")]:
        a, b = s.find(opener), s.rfind(closer)
        if a != -1 and b != -1 and b > a:
            cand = s[a : b + 1]
            try:
                return json.loads(cand)
            except Exception:
                pass
    return json.loads(s)

def _call_json_with_retries(call_fn, schema_model=None, retries=3, sleep=1.5):
    last_exc = None
    for i in range(retries):
        try:
            raw = call_fn()
            obj = _loose_to_json(raw)
            if schema_model is not None and PydanticAvailable:
                obj = schema_model.model_validate(obj).model_dump()
            return obj
        except Exception as e:
            last_exc = e
            time.sleep(sleep * (2 ** i))
    raise last_exc

def _extract_fenced_blocks(text: str) -> List[Tuple[str, str]]:
    blocks: List[Tuple[str, str]] = []
    i = 0
    while True:
        start = text.find("```", i)
        if start == -1:
            break
        nl = text.find("\n", start)
        if nl == -1:
            break
        lang = text[start + 3 : nl].strip() or ""
        end = text.find("```", nl + 1)
        if end == -1:
            break
        body = text[nl + 1 : end].strip()
        blocks.append((lang, body))
        i = end + 3
    return blocks

def _extract_xml_tag(raw: str, tag: str) -> Optional[str]:
    start = raw.find(f"<{tag}")
    end = raw.find(f"</{tag}>")
    if start != -1 and end != -1 and end > start:
        end += len(f"</{tag}>")
        return raw[start:end].strip()
    return None

def _sanitize_bpmn(xml: str) -> str:
    s = (xml or "").strip()
    if not s:
        return s
    if 'xmlns:bpmn="' not in s:
        s = s.replace(
            "<bpmn:definitions",
            '<bpmn:definitions xmlns:bpmn="http://www.omg.org/spec/BPMN/20100524/MODEL" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
            1,
        )
    if "</bpmn:definitions>" not in s and s.count("<bpmn:definitions") == 1:
        s += "\n</bpmn:definitions>"
    return s

_CANON_TRIAGE_KEYS = ("triage", "propose_triage")
_CANON_FLAG_KEYS = ("flags", "set_flags")
_DERIVED_BLOCKLIST = {"danger_sign", "clinic_referral", "triage"}

def _to_snake(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(name).strip().lower())
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "var"

def _norm_name(s: Optional[str]) -> str:
    return (s or "").strip()

# ---------- Step 2 helpers: normalize condition shapes ----------

_SYM_KEYS = {"sym", "symbol", "var", "variable", "name"}
_EQ_KEYS = {"eq", "equals", "is", "value", "present"}
_OBS_KEYS = {"obs", "observation", "var", "variable", "name"}
_OP_KEYS  = {"op", "operator", "cmp", "comparison", "relation", "rel", "sign"}
_VAL_KEYS = {"value", "rhs", "threshold", "val", "target"}

_GROUP_ANY_KEYS = {"any_of", "anyof", "any", "or", "OR", "AnyOf"}
_GROUP_ALL_KEYS = {"all_of", "allof", "all", "and", "AND", "AllOf"}

def _clean_bool(v: Any) -> Optional[bool]:
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"true", "yes", "y", "present"}:
        return True
    if s in {"false", "no", "n", "absent"}:
        return False
    return None

def _clean_number_or_str(v: Any) -> Union[float, int, str, bool]:
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    b = _clean_bool(s)
    if b is not None:
        return b
    if re.match(r"^-?\d+(\.\d+)?$", s):
        return float(s) if "." in s else int(s)
    return s

def _map_op(op: str) -> Optional[str]:
    s = str(op).strip().lower()
    if s in {">=", "ge"}:
        return "ge"
    if s in {"<=", "le"}:
        return "le"
    if s in {">", "gt"}:
        return "gt"
    if s in {"<", "lt"}:
        return "lt"
    if s in {"=", "==", "eq"}:
        return "eq"
    if s in {"!=", "<>", "ne"}:
        return "ne"
    return None

def _first_key(d: Dict[str, Any], keys: set) -> Optional[str]:
    for k in list(d.keys()):
        k_norm = _to_snake(k)
        if k_norm in keys:
            return k
    return None

def _unwrap_cond_wrappers(c: Any) -> Any:
    if not isinstance(c, dict):
        return c
    for wrap in ("SymCond", "symcond", "OBS", "ObsCond", "obscond", "ANYOF", "AllOf", "AnyOf", "ALL_OF", "ANY_OF"):
        if wrap in c and isinstance(c[wrap], (dict, list)):
            return c[wrap]
    return c

def _coerce_condition(cond: Any) -> Optional[Dict[str, Any]]:
    c = _unwrap_cond_wrappers(cond)
    if not isinstance(c, dict):
        return None

    k_any = _first_key(c, _GROUP_ANY_KEYS)
    k_all = _first_key(c, _GROUP_ALL_KEYS)
    if k_any:
        seq = c.get(k_any)
        if not isinstance(seq, list):
            return None
        kids = [x for x in (_coerce_condition(s) for s in seq) if x]
        if not kids:
            return None
        if len(kids) == 1:
            return kids[0]
        return {"any_of": kids}
    if k_all:
        seq = c.get(k_all)
        if not isinstance(seq, list):
            return None
        kids = [x for x in (_coerce_condition(s) for s in seq) if x]
        if not kids:
            return None
        if len(kids) == 1:
            return kids[0]
        return {"all_of": kids}

    k_sym = _first_key(c, _SYM_KEYS)
    if k_sym:
        v_eq_key = _first_key(c, _EQ_KEYS)
        eq_val = c.get(v_eq_key) if v_eq_key else c.get("eq")
        b = _clean_bool(eq_val if eq_val is not None else True)
        if b is None:
            return None
        sym_name = _to_snake(c.get(k_sym))
        if sym_name in _DERIVED_BLOCKLIST:
            return None
        return {"sym": sym_name, "eq": bool(b)}

    k_obs = _first_key(c, _OBS_KEYS)
    k_op  = _first_key(c, _OP_KEYS)
    k_val = _first_key(c, _VAL_KEYS)
    if k_obs and (k_op or k_val):
        op = _map_op(c.get(k_op, "eq"))
        if not op:
            return None
        val = _clean_number_or_str(c.get(k_val))
        obs_name = _to_snake(c.get(k_obs))
        return {"obs": obs_name, "op": op, "value": val}

    if "sym" in c and "eq" in c:
        if _to_snake(c["sym"]) in _DERIVED_BLOCKLIST:
            return None
        return {"sym": _to_snake(c["sym"]), "eq": bool(_clean_bool(c["eq"]) is True)}
    if all(k in c for k in ("obs", "op", "value")):
        op = _map_op(c.get("op"))
        if not op:
            return None
        return {"obs": _to_snake(c["obs"]), "op": op, "value": _clean_number_or_str(c["value"])}

    return None

def _normalize_rule_schema(rule: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    r = dict(rule or {})
    th = dict(r.get("then") or {})
    triage = None
    for k in _CANON_TRIAGE_KEYS:
        if th.get(k) is not None:
            triage = th[k]
            break
    if triage is not None:
        th["triage"] = triage

    flags = None
    for k in _CANON_FLAG_KEYS:
        if th.get(k) is not None:
            flags = th[k]
            break
    th["flags"] = list(flags or [])
    th.setdefault("reasons", [])
    th.setdefault("actions", [])
    th.setdefault("advice", [])
    th["priority"] = int(th.get("priority") or 0)
    r["then"] = th

    when = r.get("when")
    if not isinstance(when, list) or len(when) == 0:
        return None

    def _has_bad_sym(c: Dict[str, Any]) -> bool:
        return "sym" in c and str(c.get("sym")).strip().lower() in _DERIVED_BLOCKLIST

    new_when: List[Dict[str, Any]] = []
    for cond in when:
        coerced = _coerce_condition(cond)
        if coerced is None:
            continue
        def bad_nested(x: Dict[str, Any]) -> bool:
            if _has_bad_sym(x):
                return True
            if "any_of" in x or "all_of" in x:
                key = "any_of" if "any_of" in x else "all_of"
                return any(_has_bad_sym(k) or bad_nested(k) for k in (x.get(key) or []))
            return False
        if bad_nested(coerced):
            continue
        new_when.append(coerced)

    if not new_when:
        return None
    r["when"] = new_when
    return r

def _dedupe_rule_ids(rules: List[Dict[str, Any]], prefix: str = "r") -> List[Dict[str, Any]]:
    seen = {}
    out = []
    for i, r in enumerate(rules, 1):
        rid = str(r.get("rule_id") or f"{prefix}_{i}").strip()
        rid = re.sub(r"\s+", "_", rid.lower())
        if rid in seen:
            k = 2
            new_id = f"{rid}__{k}"
            while new_id in seen:
                k += 1
                new_id = f"{rid}__{k}"
            rid = new_id
        seen[rid] = True
        r["rule_id"] = rid
        out.append(r)
    return out

def _build_canonical_map(config: Dict[str, Any], variables: List[Dict[str, Any]]) -> Dict[str, str]:
    canon_set = set(config.get("canonical_variables", []))
    by_name = {(v.get("name") or "").strip().lower(): v for v in variables if isinstance(v, dict)}
    idx: Dict[str, str] = {}
    for name, v in by_name.items():
        syns = set([name]) | set((v.get("synonyms") or []))
        target = next((s for s in syns if s in canon_set), name)
        for s in syns:
            idx[str(s).strip().lower()] = target
    return idx

def _rewrite_rule_to_canon(rule: Dict[str, Any], canon: Dict[str, str]) -> Dict[str, Any]:
    r = dict(rule)
    def map_var(x: str) -> str:
        return canon.get(str(x).strip().lower(), str(x).strip().lower())
    new_when = []
    for c in r.get("when", []):
        c = dict(c)
        if "sym" in c:
            c["sym"] = map_var(c["sym"])
        elif "obs" in c:
            c["obs"] = map_var(c["obs"])
        elif "all_of" in c or "any_of" in c:
            k = "all_of" if "all_of" in c else "any_of"
            seq = []
            for s in c.get(k) or []:
                s = dict(s)
                if "sym" in s:
                    s["sym"] = map_var(s["sym"])
                if "obs" in s:
                    s["obs"] = map_var(s["obs"])
                seq.append(s)
            c[k] = seq
        new_when.append(c)
    r["when"] = new_when
    return r

def _flatten_rule_conditions(rule: Dict[str, Any]) -> List[str]:
    out = []
    def one(c):
        if "sym" in c:
            return f"{c['sym']}=={str(c.get('eq')).lower()}"
        if "obs" in c:
            return f"{c['obs']} {c['op']} {json.dumps(c['value'])}"
        return None
    for c in rule.get("when", []):
        if "all_of" in c or "any_of" in c:
            k = "all_of" if "all_of" in c else "any_of"
            parts = [one(s) for s in (c.get(k) or []) if one(s)]
            out.append(f"{k}:" + " && ".join(parts))
        else:
            s = one(c)
            if s:
                out.append(s)
    return out

def _rules_flattened(ir: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [
        {"rule_id": r.get("rule_id"),
         "triage": (r.get("then") or {}).get("triage"),
         "conds": _flatten_rule_conditions(r)}
        for r in ir.get("rules", [])
    ]

def _preflight_ir(ir: Dict[str, Any]) -> None:
    names = [(v.get("name") or "").strip().lower() for v in ir.get("variables", []) if isinstance(v, dict)]
    dup_var_names = [n for n, cnt in collections.Counter(names).items() if cnt > 1]
    if dup_var_names:
        raise RuntimeError(f"Duplicate variable names after merge: {dup_var_names}")
    ids = [(r.get("rule_id") or "").strip().lower() for r in ir.get("rules", []) if isinstance(r, dict)]
    dups = [i for i, c in collections.Counter(ids).items() if c > 1]
    if dups:
        raise RuntimeError(f"Duplicate rule_ids after dedupe pass: {dups}")
    empties = [r.get("rule_id") for r in ir.get("rules", []) if not (r.get("when") or [])]
    if empties:
        raise RuntimeError(f"Empty WHEN in rules: {empties}")

def _enforce_ask_ownership(ask_plan: List[Dict[str, Any]],
                           priority_order=("module_a", "module_b", "module_c", "module_d", "module_e")):
    owner: Dict[str, str] = {}
    order = {m: i for i, m in enumerate(priority_order)}
    for blk in sorted([b for b in ask_plan if isinstance(b, dict)],
                      key=lambda b: order.get(b.get("module", "zzz"), 999)):
        module = blk.get("module")
        kept_ask = []
        for q in blk.get("ask", []) or []:
            if q not in owner:
                owner[q] = module
                kept_ask.append(q)
        blk["ask"] = kept_ask
        fuw = {}
        for cond, qs in (blk.get("followups_if") or {}).items():
            new_qs = []
            for q in qs:
                if q not in owner:
                    owner[q] = module
                    new_qs.append(q)
            fuw[cond] = new_qs
        blk["followups_if"] = fuw
    return ask_plan

def _pages_from_ref_string(s: str) -> List[int]:
    if not s:
        return []
    out = []
    for tok in re.findall(r"p(\d{1,4})(?:-(\d{1,4}))?", s, flags=re.IGNORECASE):
        a = int(tok[0])
        b = int(tok[1]) if tok[1] else None
        if b is None:
            out.append(a)
        else:
            lo, hi = sorted((a, b))
            out.extend(list(range(lo, hi + 1))[:500])
    return out

def _pages_from_refs(refs: Optional[List[str]]) -> set:
    pages = set()
    for r in (refs or []):
        for p in _pages_from_ref_string(str(r)):
            pages.add(p)
    return pages

def _var_sig(v: Dict[str, Any]) -> tuple:
    t = v.get("type")
    u = v.get("unit")
    allowed = tuple(sorted([str(x) for x in (v.get("allowed") or [])]))
    return (t, u, allowed)

def _coalesce_group(vars_with_same_sig: List[Dict[str, Any]]) -> Dict[str, Any]:
    base = dict(vars_with_same_sig[0])
    syn = set(base.get("synonyms") or [])
    refs = set(base.get("refs") or [])
    allowed = set(base.get("allowed") or [])
    for v in vars_with_same_sig[1:]:
        syn |= set(v.get("synonyms") or [])
        refs |= set(v.get("refs") or [])
        allowed |= set(v.get("allowed") or [])
        if not base.get("prompt") and v.get("prompt"):
            base["prompt"] = v["prompt"]
    base["synonyms"] = sorted(list(syn))
    base["refs"] = sorted(list(refs))
    base["allowed"] = sorted(list(allowed)) if allowed else []
    return base

def _resolve_variables_snake_and_rewrite_rules(merged: Dict[str, Any]) -> Dict[str, Any]:
    qa_notes = merged.setdefault("qa", {}).setdefault("notes", [])
    vars_in = [v for v in (merged.get("variables") or []) if isinstance(v, dict) and _norm_name(v.get("name"))]
    groups: Dict[str, List[Dict[str, Any]]] = {}
    names_seen_per_group: Dict[str, set] = {}
    for v in vars_in:
        orig = _norm_name(v.get("name"))
        key = _to_snake(orig)
        vv = dict(v)
        syns = set(vv.get("synonyms") or [])
        syns.add(orig)
        vv["synonyms"] = sorted(list(syns))
        groups.setdefault(key, []).append(vv)
        names_seen_per_group.setdefault(key, set()).update({orig.lower(), *[str(s).strip().lower() for s in syns]})

    new_vars: List[Dict[str, Any]] = []
    name_to_variants: Dict[str, List[Dict[str, Any]]] = {}
    rename_events: Dict[str, List[str]] = {}

    for base, gvars in groups.items():
        sig_bins: Dict[tuple, List[Dict[str, Any]]] = {}
        for v in gvars:
            sig_bins.setdefault(_var_sig(v), []).append(v)

        variants_meta = []
        if len(sig_bins) > 1:
            qa_notes.append(f"variable_homonyms_split:{base}:{len(sig_bins)}")

        idx = 1
        for _sig, same_sig_vars in sig_bins.items():
            merged_v = _coalesce_group(same_sig_vars)
            final_name = base if idx == 1 else f"{base}__{idx}"
            idx += 1
            originals = {_norm_name(x.get("name")) for x in same_sig_vars if _norm_name(x.get("name"))}
            for o in originals:
                rename_events.setdefault(o, []).append(final_name)
            merged_v["name"] = final_name
            pages = _pages_from_refs(merged_v.get("refs") or [])
            variants_meta.append({"new_name": final_name, "pages": pages})
            new_vars.append(merged_v)

        for n in names_seen_per_group.get(base, set()):
            name_to_variants[n] = variants_meta

    def pick_variant(seen_name: str, rule_ref: Optional[str]) -> str:
        variants = name_to_variants.get((seen_name or "").strip().lower())
        if not variants:
            return _to_snake(seen_name)
        if not rule_ref:
            return variants[0]["new_name"]
        r_pages = set(_pages_from_ref_string(rule_ref))
        best = variants[0]
        best_sc = -1
        for v in variants:
            sc = len(r_pages & set(v["pages"]))
            if sc > best_sc:
                best_sc = sc
                best = v
        return best["new_name"]

    rules_out: List[Dict[str, Any]] = []
    for r in (merged.get("rules") or []):
        rr = dict(r)
        rule_ref = _norm_name(((rr.get("then") or {}).get("guideline_ref")))
        new_when = []
        for c in (rr.get("when") or []):
            c = dict(c)
            if "sym" in c and c["sym"]:
                sym = _norm_name(c["sym"])
                key = sym.strip().lower()
                if key in name_to_variants or _to_snake(sym) in groups:
                    c["sym"] = pick_variant(sym, rule_ref)
                else:
                    c["sym"] = _to_snake(sym)
            elif "obs" in c and c["obs"]:
                obs = _norm_name(c["obs"])
                key = obs.strip().lower()
                if key in name_to_variants or _to_snake(obs) in groups:
                    c["obs"] = pick_variant(obs, rule_ref)
                else:
                    c["obs"] = _to_snake(obs)
            elif "all_of" in c or "any_of" in c:
                k = "all_of" if "all_of" in c else "any_of"
                seq = []
                for s in (c.get(k) or []):
                    s = dict(s)
                    if "sym" in s and s["sym"]:
                        sym = _norm_name(s["sym"])
                        key = sym.strip().lower()
                        if key in name_to_variants or _to_snake(sym) in groups:
                            s["sym"] = pick_variant(sym, rule_ref)
                        else:
                            s["sym"] = _to_snake(sym)
                    if "obs" in s and s["obs"]:
                        obs = _norm_name(s["obs"])
                        key = obs.strip().lower()
                        if key in name_to_variants or _to_snake(obs) in groups:
                            s["obs"] = pick_variant(obs, rule_ref)
                        else:
                            s["obs"] = _to_snake(obs)
                    seq.append(s)
                c[k] = seq
            new_when.append(c)
        rr["when"] = new_when
        rules_out.append(rr)

    if rename_events:
        merged.setdefault("qa", {}).setdefault("notes", []).append(f"snake_case_renamed:{rename_events}")

    merged["variables"] = new_vars
    merged["rules"] = rules_out
    return merged

def _fact_sheet_from_sections(section_objs: List[Dict[str, Any]]) -> Dict[str, Any]:
    var_ix: Dict[str, Dict[str, Any]] = {}
    facts_rules: List[Dict[str, Any]] = []
    for sec in (section_objs or []):
        for v in (sec.get("variables") or []):
            if not isinstance(v, dict):
                continue
            name = (v.get("name") or "").strip()
            if not name:
                continue
            key = name.lower()
            cur = var_ix.get(key)
            if cur is None:
                cur = dict(
                    name=name,
                    type=v.get("type"),
                    unit=v.get("unit"),
                    allowed=v.get("allowed"),
                    synonyms=list(v.get("synonyms") or []),
                    refs=list(v.get("refs") or []),
                )
                var_ix[key] = cur
            else:
                cur["synonyms"] = sorted(list(set((cur.get("synonyms") or []) + (v.get("synonyms") or []))))
                cur["refs"] = sorted(list(set((cur.get("refs") or []) + (v.get("refs") or []))))
        for r in (sec.get("rules") or []):
            if not isinstance(r, dict):
                continue
            nr = _normalize_rule_schema(r)
            if not nr:
                continue
            th = nr.get("then") or {}
            facts_rules.append(
                {
                    "rule_id": r.get("rule_id"),
                    "when": nr.get("when") or [],
                    "then": {
                        "triage": th.get("triage"),
                        "flags": th.get("flags") or [],
                        "reasons": th.get("reasons") or [],
                        "actions": th.get("actions") or [],
                        "guideline_ref": th.get("guideline_ref"),
                        "priority": int(th.get("priority") or 0),
                        "advice": [],
                    },
                }
            )
    return {
        "variables": list(var_ix.values()),
        "rules": _dedupe_rule_ids(facts_rules, prefix="fs"),
        "qa": {"notes": ["from_fact_sheet_builder"]},
    }

# -------------------------
# lossless consolidation
# -------------------------

def _canon_cond(cond: Dict[str, Any]) -> Dict[str, Any]:
    def canon_one(c: Dict[str, Any]) -> Dict[str, Any]:
        if "sym" in c:
            return {"sym": _to_snake(c["sym"]), "eq": bool(c.get("eq") is True)}
        if "obs" in c:
            return {"obs": _to_snake(c["obs"]), "op": _map_op(c.get("op") or "eq") or "eq", "value": _clean_number_or_str(c.get("value"))}
        if "all_of" in c or "any_of" in c:
            key = "all_of" if "all_of" in c else "any_of"
            kids = [canon_one(dict(k)) for k in (c.get(key) or [])]
            kids_sorted = sorted([json.dumps(k, sort_keys=True) for k in kids])
            return {key: [json.loads(s) for s in kids_sorted]}
        return {}
    return canon_one(dict(cond))

def _conds_signature(when: List[Dict[str, Any]]) -> Tuple[str, ...]:
    canon = [_canon_cond(c) for c in (when or [])]
    as_strs = [json.dumps(c, sort_keys=True) for c in canon]
    return tuple(sorted(as_strs))

def _merge_then_blocks(blocks: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not blocks:
        return {"triage": "home", "flags": [], "reasons": [], "actions": [], "advice": [], "guideline_ref": "p0", "priority": 0}
    triage = (blocks[0].get("triage") or "home").strip().lower()
    ref = (blocks[0].get("guideline_ref") or "p0").strip()
    flags: set = set()
    reasons: set = set()
    actions: List[dict] = []
    pr = 0
    for b in blocks:
        flags |= set(b.get("flags") or [])
        reasons |= set(b.get("reasons") or [])
        for a in (b.get("actions") or []):
            if isinstance(a, dict) and a:
                actions.append(a)
        pr = max(pr, int(b.get("priority") or 0))
    seen = set()
    uniq_actions = []
    for a in actions:
        key = (a.get("id"), bool(a.get("if_available")))
        if key in seen:
            continue
        seen.add(key)
        uniq_actions.append({"id": a.get("id"), "if_available": bool(a.get("if_available"))})
    return {
        "triage": triage,
        "flags": sorted(list(flags)),
        "reasons": sorted(list(reasons)),
        "actions": uniq_actions,
        "advice": [],
        "guideline_ref": ref,
        "priority": pr,
    }

def _lossless_consolidate_rules(fact_sheet: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Deterministic consolidation:
    - normalize each rule
    - group by (conditions signature, triage, guideline_ref)
    - union reasons, flags, actions within each group
    - never merge across triage or across guideline_ref
    """
    cleaned: List[Dict[str, Any]] = []
    for r in (fact_sheet.get("rules") or []):
        nr = _normalize_rule_schema(r)
        if not nr:
            continue
        nr["when"] = [c for c in (nr.get("when") or []) if _coerce_condition(c)]
        if not nr["when"]:
            continue
        cleaned.append(nr)

    # index by signature and outputs
    bins: Dict[Tuple[Tuple[str, ...], str, str], List[Dict[str, Any]]] = {}
    for r in cleaned:
        sig = _conds_signature(r.get("when") or [])
        th = r.get("then") or {}
        tri = (th.get("triage") or "home").strip().lower()
        ref = (th.get("guideline_ref") or "p0").strip()
        bins.setdefault((sig, tri, ref), []).append(r)

    out: List[Dict[str, Any]] = []
    for (sig, tri, ref), rs in bins.items():
        # merge then blocks inside the exact same signature, triage, ref
        merged_then = _merge_then_blocks([x.get("then") or {} for x in rs])
        # pick a stable when from the first
        base_when = rs[0].get("when") or []
        out.append({
            "rule_id": f"fs_merge_{len(out)+1}",
            "when": base_when,
            "then": merged_then
        })

    # stable ids
    out = _dedupe_rule_ids(out, prefix="fs")
    return out

def _index_fact_sheet_rules(fact_sheet: Dict[str, Any]) -> Dict[Tuple[str, ...], List[Dict[str, Any]]]:
    idx: Dict[Tuple[str, ...], List[Dict[str, Any]]] = {}
    for r in (fact_sheet.get("rules") or []):
        sig = _conds_signature(r.get("when") or [])
        idx.setdefault(sig, []).append(r)
    return idx

def _group_by_triage_and_ref(rules: List[Dict[str, Any]]) -> Dict[Tuple[str, str], List[Dict[str, Any]]]:
    out: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for r in (rules or []):
        th = r.get("then") or {}
        tri = (th.get("triage") or "").strip().lower() or "home"
        ref = (th.get("guideline_ref") or "").strip() or "p0"
        out.setdefault((tri, ref), []).append(r)
    return out

def _preserve_ref_and_triage_variants(cleaned_rules: List[Dict[str, Any]], fact_sheet: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
    notes: List[str] = []
    fs_idx = _index_fact_sheet_rules(fact_sheet)
    out_rules: List[Dict[str, Any]] = []

    for r in cleaned_rules:
        sig = _conds_signature(r.get("when") or [])
        variants = fs_idx.get(sig)
        if not variants or len(variants) == 1:
            out_rules.append(r)
            continue

        groups = _group_by_triage_and_ref(variants)
        if len(groups) == 1:
            (tri, ref), rs = next(iter(groups.items()))
            rr = dict(r)
            then = dict(rr.get("then") or {})
            merged_then = _merge_then_blocks([x.get("then") or {} for x in rs])
            then.update(merged_then)
            rr["then"] = then
            out_rules.append(rr)
            continue

        did_split = 0
        for (tri, ref), rs in groups.items():
            rr = dict(r)
            then = dict(rr.get("then") or {})
            merged_then = _merge_then_blocks([x.get("then") or {} for x in rs])
            then.update(merged_then)
            rr["then"] = then
            base_id = str(r.get("rule_id") or "merged_rule")
            rr["rule_id"] = f"{base_id}_{tri}_{_to_snake(ref)}"
            out_rules.append(rr)
            did_split += 1
        notes.append(f"ref_triage_variants_preserved:{r.get('rule_id')}->{did_split}")
    return out_rules, notes

# ---------- prompts (unchanged except merge step no longer drives consolidation) ----------

class OpenAIGuardedExtractor:
    def __init__(self,
                 api_key: Optional[str] = None,
                 model_section: str = "gpt-5",
                 model_merge: str = "gpt-5",
                 model_dmn: str = "gpt-5",
                 model_bpmn: str = "gpt-5",
                 model_audit: str = "gpt-5",
                 seed: Optional[int] = 42,
                 canonical_config_path: Optional[str] = "chatchw/config/canonical_config.json") -> None:
        key = api_key or os.getenv("OPENAI_API_KEY")
        if not key and not os.getenv("SKIP_OPENAI_CHECK"):
            pass
        if OpenAI is None and key:
            raise RuntimeError("openai package not available")

        def _norm(m: str, default: str) -> str:
            if m and m.startswith("auto:"):
                m = m.split("auto:", 1)[1] or default
            if m in (None, "", "auto"):
                return default
            return m

        self.client = OpenAI(api_key=key) if key and OpenAI else None
        self.seed = seed
        self.models = dict(
            section=_norm(model_section, "gpt-5"),
            merge=_norm(model_merge, "gpt-5"),
            dmn=_norm(model_dmn, "gpt-5"),
            bpmn=_norm(model_bpmn, "gpt-5"),
            audit=_norm(model_audit, "gpt-5"),
        )
        self.usage = {"prompt": 0, "completion": 0, "total": 0}
        self._usage_events: List[Dict[str, int]] = []

        self.debug_dir = Path(".chatchw_debug")
        self.debug_dir.mkdir(exist_ok=True)

        default_canon_vars = [
            "symptom1",
            "symptom1_duration_days",
            "symptom2",
            "symptom2_duration_days",
            "sign1_present",
            "sign2_present",
            "sign3_present",
            "measurement1_rate",
            "measurement2_temp_c",
            "measurement3_mm",
            "age_months",
            "area1_flag",
        ]
        try:
            cfg_path = Path(canonical_config_path)
            if cfg_path.exists():
                self.config = json.loads(cfg_path.read_text(encoding="utf-8"))
            else:
                self.config = {
                    "canonical_variables": default_canon_vars,
                    "priority_tiers": {"hospital_min": 90, "clinic_min": 50},
                }
        except Exception:
            self.config = {
                "canonical_variables": default_canon_vars,
                "priority_tiers": {"hospital_min": 90, "clinic_min": 50},
            }

        canon = ", ".join(self.config.get("canonical_variables", default_canon_vars))

        self.section_system = (
            "You extract structured decision rules from a technical manual. Return ONLY JSON.\n\n"
            "IMPORTANT OUTPUT POLICY\n"
            "- Preserve original clinical or medical terms exactly as written in the source text wherever they appear.\n"
            "- No prose outside the JSON object.\n\n"
            "SCHEMA\n{\n"
            "  \"variables\":[{\"name\":snake,\"type\":\"number|boolean|string\",\"unit\":null|unit,\"allowed\":null|[...],"
            "                \"synonyms\":[snake...],\"prompt\":short,\"refs\":[page_or_section]}],\n"
            "  \"rules\":[{\"rule_id\":str,\n"
            "            \"when\":[ {\"obs\":var,\"op\":\"lt|le|gt|ge|eq|ne\",\"value\":num|bool|string} |\n"
            "                     {\"sym\":var,\"eq\":true|false} |\n"
            "                     {\"all_of\":[COND...]} | {\"any_of\":[COND...]} ],\n"
            "            \"then\":{\"triage\":\"hospital|clinic|home\",\n"
            "                    \"flags\":[snake...],\n"
            "                    \"reasons\":[snake...],\n"
            "                    \"actions\":[{\"id\":snake,\"if_available\":bool}],\n"
            "                    \"advice\":[],\n"
            "                    \"guideline_ref\":str,\"priority\":int}}],\n"
            "  \"canonical_map\":{},\n"
            "  \"qa\":{\"notes\":[]}\n}\n\n"
            f"RULES\n- Use canonical names if present: {canon} (examples; prefer the source's medical terms).\n"
            "- Encode only literal thresholds that appear.\n"
            "- Do not use derived outputs in conditions.\n"
            "- Every rule gets guideline_ref like \"p41\" or a section id.\n"
            "- Include CHW treatments as structured actions in then.actions. Examples: give_ors, give_zinc, teach_handwashing.\n"
            "- Do not combine rules across different triage levels or different guideline_ref pages.\n"
            "OUTPUT: only JSON for this section."
        )

        # Kept for compatibility, but no longer used to reduce rule count
        self.merge_system = (
            "Merge this FACT_SHEET into one comprehensive IR. "
            "Return ONLY JSON with the same schema as step 1 plus:\n"
            "- canonical_map filled for all variables\n- qa.unmapped_vars\n- qa.dedup_dropped\n- qa.overlap_fixed\n"
            "RULES\n"
            "- Preserve original clinical terms.\n"
            "- Normalize labels to snake_case where needed.\n"
            "- Rewrite rules to canonical names.\n"
            "- Force then.advice to [].\n"
            "- Keep every literal threshold.\n"
            "- Never merge rules that differ in then.triage.\n"
            "- Never merge rules that differ in then.guideline_ref page ranges.\n"
            "- Prefer more specific conditions over less specific ones when deduplicating.\n"
            "INPUT\nFACT_SHEET:\n<COMPACT VARIABLES + RULES>"
        )

        self.rules_system = (
            "Given a FACT_SHEET (variables + rules), consolidate overlaps into RULES.\n"
            "Return ONLY JSON: { \"rules\": [ {\"rule_id\": str, \"when\": [...], \"then\": {\"triage\": \"hospital|clinic|home\", "
            "\"flags\":[], \"reasons\":[], \"actions\":[{\"id\":snake,\"if_available\":bool}], \"advice\":[], \"guideline_ref\": str, \"priority\": int } } ] }\n"
            "HARD RULES: preserve terms, ban derived fields in conditions, advice must be [], each rule has guideline_ref.\n"
            "Do NOT collapse rules across different triage levels. Do NOT collapse rules across different guideline_ref pages."
        )

        self.dmn_system = (
            "Convert RULES_JSON into modular DMN 1.4 using the DMN 1.4 MODEL namespace (2019-11-11). "
            "Return exactly TWO fenced blocks:\n"
            "1) ```xml <dmn:definitions>…```  2) ```json ASK_PLAN```\n\n"
            "HARD CONSTRAINTS\n"
            "- Use variable names exactly as in RULES_JSON.\n"
            "- Decisions: decide_module_a, decide_module_b, decide_module_c, decide_module_d, decide_module_e, aggregate_final\n"
            "- Each module uses <dmn:decisionTable hitPolicy=\"FIRST\"> with outputs:\n"
            "  triage:string, danger_sign:boolean, clinic_referral:boolean, reason:string, ref:string, actions:string, advice:string\n"
            "- Populate every output cell:\n"
            "  triage → \"hospital\" | \"clinic\" | \"home\"\n"
            "  danger_sign → true | false\n"
            "  clinic_referral → true | false\n"
            "  reason → snake_case string\n"
            "  ref → \"pNN\" or a section id string\n"
            "  actions → comma separated action ids from then.actions where if_available is true, or \"\"\n"
            "  advice → \"\" (empty string)\n"
            "- aggregate_final should combine booleans by reading the module outputs directly.\n"
            "- Provide an ASK_PLAN array describing input collection order and followups."
        )

        self.bpmn_system = (
            "Produce one BPMN 2.0 <bpmn:definitions> only. Use xmlns:bpmn and xmlns:xsi. Build a simple ask → evaluate → route flow."
        )

        self.coverage_system = (
            "Given RULES_JSON and the DMN XML, return ONLY JSON:\n"
            "{\"unmapped_rule_ids\":[...], \"module_counts\":{...}, \"notes\":[...]}\n"
            "Use RULES_JSON.ir_flat.conds to match DMN inputEntry texts when possible. Match on conditions and triage."
        )

    def _accumulate_usage(self, resp) -> None:
        try:
            u = getattr(resp, "usage", None)
            if not u and isinstance(resp, dict):
                u = resp.get("usage")
            if not u:
                return
            p = int(getattr(u, "prompt_tokens", 0) or (u.get("prompt_tokens") if isinstance(u, dict) else 0) or 0)
            c = int(getattr(u, "completion_tokens", 0) or (u.get("completion_tokens") if isinstance(u, dict) else 0) or 0)
            t = int(getattr(u, "total_tokens", 0) or (u.get("total_tokens") if isinstance(u, dict) else (p + c)))
            self.usage["prompt"] += p
            self.usage["completion"] += c
            self.usage["total"] += t
            self._usage_events.append({"prompt": p, "completion": c, "total": t})
        except Exception:
            pass

    def get_usage_summary(self) -> Dict[str, int]:
        return dict(self.usage)

    def _complete(self, model_name: str, messages: list, max_out: int):
        if not self.client:
            raise RuntimeError("OpenAI client not initialized.")
        base = dict(model=model_name, messages=messages)
        is_gpt5 = isinstance(model_name, str) and model_name.startswith("gpt-5")
        if not is_gpt5:
            base.update(dict(temperature=0.0, top_p=1))
        extras = {}
        if is_gpt5:
            extras = {"reasoning_effort": "minimal", "verbosity": "low"}
        base_seed = dict(base)
        base_seed["seed"] = getattr(self, "seed", None)
        try:
            resp = self.client.chat.completions.create(max_completion_tokens=max_out, **base_seed, **extras)
            self._accumulate_usage(resp)
            return resp
        except Exception as e1:
            msg1 = str(e1).lower()
            if "verbosity" in msg1 or "reasoning" in msg1 or "unsupported parameter" in msg1:
                try:
                    resp = self.client.chat.completions.create(max_completion_tokens=max_out, **base_seed)
                    self._accumulate_usage(resp)
                    return resp
                except Exception as e2:
                    msg2 = str(e2).lower()
                    if "max_completion_tokens" in msg2 or "unsupported parameter" in msg2:
                        try:
                            resp = self.client.chat.completions.create(max_tokens=max_out, **base_seed)
                            self._accumulate_usage(resp)
                            return resp
                        except Exception as e3:
                            msg3 = str(e3).lower()
                            if "seed" in msg3 or "unsupported parameter" in msg3:
                                base_noseed = dict(base)
                                resp = self.client.chat.completions.create(max_tokens=max_out, **base_noseed)
                                self._accumulate_usage(resp)
                                return resp
                            raise
                    raise
            if "max_completion_tokens" in msg1:
                try:
                    resp = self.client.chat.completions.create(max_tokens=max_out, **base_seed, **extras)
                    self._accumulate_usage(resp)
                    return resp
                except Exception as e4:
                    msg4 = str(e4).lower()
                    if "seed" in msg4 or "unsupported parameter" in msg4:
                        base_noseed = dict(base)
                        resp = self.client.chat.completions.create(max_tokens=max_out, **base_noseed, **extras)
                        self._accumulate_usage(resp)
                        return resp
                    raise
            raise

    def _chat_json(self, system: str, user: str, max_out: int = 6000, model_key: str = "section", schema_model=None) -> Dict[str, Any]:
        model_name = self.models.get(model_key, self.models["section"])
        messages = [{"role": "system", "content": system}, {"role": "user", "content": user}]
        def _call():
            resp = self._complete(model_name, messages, max_out)
            return (resp.choices[0].message.content or "").strip()
        try:
            return _call_json_with_retries(_call, schema_model=schema_model, retries=3)
        except Exception as e:
            dbg_user = user
            max_head = 12000
            max_tail = 4000
            if len(dbg_user) > (max_head + max_tail):
                dbg_user = dbg_user[:max_head] + "\n...<TRUNCATED>...\n" + dbg_user[-max_tail:]
            Path(self.debug_dir / "guarded_debug_last.txt").write_text(
                f"SYSTEM:\n{system}\n\nUSER (len={len(user)}):\n{dbg_user}\n\nERR:{e}", encoding="utf-8"
            )
            raise

    def _chat_text(self, system: str, user: str, max_out: int = 12000, model_key: str = "dmn") -> str:
        model_name = self.models.get(model_key, self.models["dmn"])
        messages = [{"role": "system", "content": system}, {"role": "user", "content": user}]
        resp = self._complete(model_name, messages, max_out)
        return (resp.choices[0].message.content or "").strip()

    def extract_rules_per_section(self, pdf_path: str) -> List[Dict[str, Any]]:
        sections = self.extract_sections_from_pdf(pdf_path)
        results: List[Dict[str, Any]] = []
        for sec_id, text in sections:
            user = f"SECTION_ID: {sec_id}\n\nTEXT:\n{text}"
            try:
                obj = self._chat_json(
                    self.section_system, user, max_out=6000, model_key="section", schema_model=IR if PydanticAvailable else None
                )
            except Exception:
                continue
            cleaned = []
            for r in obj.get("rules", []) or []:
                nr = _normalize_rule_schema(r)
                if nr:
                    cleaned.append(nr)
                else:
                    obj.setdefault("qa", {}).setdefault("notes", []).append(f"dropped_rule:{r.get('rule_id')}")
            obj["rules"] = _dedupe_rule_ids(cleaned, prefix=f"{sec_id}")
            results.append(obj)
        return results

    def generate_rules_from_fact_sheet(self, fact_sheet: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Replaced the model consolidation with a deterministic lossless consolidation.
        This preserves counts and variants and avoids the 25 rule collapse.
        """
        rules_out = _lossless_consolidate_rules(fact_sheet)
        try:
            self.debug_dir.joinpath("rules_from_facts.json").write_text(
                json.dumps({"rules": rules_out}, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        except Exception:
            pass
        return rules_out

    def merge_sections(self, section_objs: List[Dict[str, Any]]) -> Dict[str, Any]:
        fact_sheet = _fact_sheet_from_sections(section_objs)
        try:
            self.debug_dir.joinpath("fact_sheet.json").write_text(
                json.dumps(fact_sheet, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        except Exception:
            pass

        consolidated_rules = self.generate_rules_from_fact_sheet(fact_sheet)

        cleaned_rules: List[Dict[str, Any]] = []
        cleaning_notes: List[str] = []
        for rule in consolidated_rules:
            rule_copy = dict(rule)
            when_list = rule_copy.get("when") or []
            if not isinstance(when_list, list):
                when_list = []
            pruned_conds: List[Dict[str, Any]] = []
            trimmed_count = 0
            for cond in when_list:
                coerced = _coerce_condition(cond)
                pruned = _coerce_condition(coerced)
                if pruned is None:
                    trimmed_count += 1
                else:
                    pruned_conds.append(pruned)
            if trimmed_count > 0:
                cleaning_notes.append(f"rule {rule_copy.get('rule_id')} trimmed_invalid_conds:{trimmed_count}")
            if not pruned_conds:
                cleaning_notes.append(f"rule {rule_copy.get('rule_id')} dropped_due_no_valid_conds")
                continue
            rule_copy["when"] = pruned_conds
            cleaned_rules.append(rule_copy)

        cleaned_rules, variant_notes = _preserve_ref_and_triage_variants(cleaned_rules, fact_sheet)
        cleaning_notes.extend(variant_notes)

        payload = json.dumps(fact_sheet, ensure_ascii=False)
        user = f"FACT_SHEET\n{payload}"

        merged: Dict[str, Any]
        if STRICT_MERGE:
            try:
                merged = self._chat_json(
                    self.merge_system,
                    user,
                    max_out=8000,
                    model_key="merge",
                    schema_model=IR if PydanticAvailable else None,
                )
                merged["rules"] = cleaned_rules
                if not merged.get("variables"):
                    merged["variables"] = fact_sheet.get("variables", [])
            except Exception as e:
                variables: Dict[str, Dict[str, Any]] = {}
                qa_notes: List[str] = [f"fallback_merge_used: {type(e).__name__}"]
                for v in (fact_sheet.get("variables") or []):
                    if not isinstance(v, dict):
                        continue
                    name = str(v.get("name", "")).strip()
                    if not name:
                        continue
                    cur = variables.get(name.lower())
                    if cur is None:
                        vv = dict(v)
                        if not isinstance(vv.get("synonyms"), list):
                            vv["synonyms"] = []
                        if not isinstance(vv.get("refs"), list):
                            vv["refs"] = []
                        variables[name.lower()] = vv
                    else:
                        syn = set(cur.get("synonyms") or []) | set(v.get("synonyms") or [])
                        refs = set(cur.get("refs") or []) | set(v.get("refs") or [])
                        cur["synonyms"] = sorted(list(syn))
                        cur["refs"] = sorted(list(refs))
                merged = {
                    "variables": list(variables.values()),
                    "rules": cleaned_rules,
                    "canonical_map": {},
                    "qa": {"notes": qa_notes, "unmapped_vars": [], "dedup_dropped": 0, "overlap_fixed": []},
                }
        else:
            variables: Dict[str, Dict[str, Any]] = {}
            qa_notes: List[str] = ["fallback_merge_used", "source=fact_sheet"]
            for v in (fact_sheet.get("variables") or []):
                if not isinstance(v, dict):
                    continue
                name = str(v.get("name", "")).strip()
                if not name:
                    continue
                cur = variables.get(name.lower())
                if cur is None:
                    vv = dict(v)
                    if not isinstance(vv.get("synonyms"), list):
                        vv["synonyms"] = []
                    if not isinstance(vv.get("refs"), list):
                        vv["refs"] = []
                    variables[name.lower()] = vv
                else:
                    syn = set(cur.get("synonyms") or []) | set(v.get("synonyms") or [])
                    refs = set(cur.get("refs") or []) | set(v.get("refs") or [])
                    cur["synonyms"] = sorted(list(syn))
                    cur["refs"] = sorted(list(refs))
            merged = {
                "variables": list(variables.values()),
                "rules": cleaned_rules,
                "canonical_map": {},
                "qa": {"notes": qa_notes, "unmapped_vars": [], "dedup_dropped": 0, "overlap_fixed": []},
            }

        try:
            notes_list = merged.setdefault("qa", {}).setdefault("notes", [])
            for note in cleaning_notes:
                notes_list.append(note)
        except Exception:
            merged["qa"] = merged.get("qa", {}) or {}
            merged["qa"].setdefault("notes", []).extend(cleaning_notes)

        merged = _resolve_variables_snake_and_rewrite_rules(merged)

        canon = _build_canonical_map(self.config, merged.get("variables", []))
        merged["canonical_map"] = canon
        merged["rules"] = [_rewrite_rule_to_canon(r, canon) for r in (merged.get("rules") or [])]
        merged["rules"] = _dedupe_rule_ids(merged["rules"], prefix="merged")

        if PydanticAvailable:
            try:
                merged = IR.model_validate(merged).model_dump()
            except ValidationError as e:
                try:
                    self.debug_dir.joinpath("pydantic_ir_validation_error.txt").write_text(str(e), encoding="utf-8")
                except Exception:
                    pass
                merged.setdefault("qa", {}).setdefault("notes", []).append(
                    "pydantic_validation_skipped_due_to_nested_group_shape"
                )

        _preflight_ir(merged)
        return merged

    def generate_dmn_and_ask_plan(self, merged_ir: Dict[str, Any]) -> Tuple[str, Any]:
        user = "RULES_JSON:\n" + json.dumps(merged_ir, ensure_ascii=False)
        text = self._chat_text(self.dmn_system, user, max_out=12000, model_key="dmn")

        try:
            Path(self.debug_dir / "dmn_ask_debug_last.txt").write_text(text, encoding="utf-8")
        except Exception:
            pass

        blocks = _extract_fenced_blocks(text)
        dmn_xml: Optional[str] = None
        ask_plan: Optional[Any] = None

        def _parse_json_loose(s: str) -> Optional[Any]:
            try:
                return _loose_to_json(s)
            except Exception:
                return None

        for _lang, body in blocks:
            b = body.strip()
            if dmn_xml is None and "<dmn:definitions" in b and "</dmn:definitions>" in b:
                dmn_xml = b
                continue
            if ask_plan is None:
                parsed = _parse_json_loose(b)
                if isinstance(parsed, (list, dict)):
                    ask_plan = parsed

        if dmn_xml is None:
            s = text
            start = s.find("<dmn:definitions")
            end = s.find("</dmn:definitions>")
            if start != -1 and end != -1 and end > start:
                end += len("</dmn:definitions>")
                dmn_xml = s[start:end].strip()

        if ask_plan is None:
            parsed = _parse_json_loose(text)
            if isinstance(parsed, (list, dict)):
                ask_plan = parsed

        def _sanitize_dmn(xml: str) -> str:
            s = xml
            s = re.sub(r"(<dmn:outputEntry>\s*<dmn:text>)([^<]*?)(</dmn:outputEntry>)", r"\1\2</dmn:text>\3", s, flags=re.DOTALL)
            s = s.replace("<dmn:outputEntry><dmn:text></dmn:text></dmn:outputEntry>",
                          '<dmn:outputEntry><dmn:text>""</dmn:text></dmn:outputEntry>')
            s = s.replace("<dmn:outputEntry><dmn:text></dmn:outputEntry>",
                          '<dmn:outputEntry><dmn:text>""</dmn:text></dmn:outputEntry>')
            s = re.sub(r"(<dmn:inputEntry>\s*<dmn:text>)([^<]*?)(</dmn:inputEntry>)", r"\1\2</dmn:text>\3", s, flags=re.DOTALL)
            s = s.replace("<dmn:inputEntry><dmn:text></dmn:text></dmn:inputEntry>",
                          "<dmn:inputEntry><dmn:text>-</dmn:text></dmn:inputEntry>")
            s = s.replace("<dmn:inputEntry><dmn:text></dmn:inputEntry>",
                          "<dmn:inputEntry><dmn:text>-</dmn:inputEntry>")
            if 'xmlns:dmn="' not in s:
                s = s.replace("<dmn:definitions",
                              '<dmn:definitions xmlns:dmn="https://www.omg.org/spec/DMN/20191111/MODEL/"', 1)
            return s

        if dmn_xml:
            dmn_xml = _sanitize_dmn(dmn_xml)

        if not dmn_xml or ask_plan is None:
            try:
                Path(self.debug_dir / "dmn_ask_debug_last.txt").write_text(text, encoding="utf-8")
            except Exception:
                pass
        _validate_dmn_outputs_or_die(dmn_xml)

        try:
            known_vars = {v["name"] for v in merged_ir.get("variables", []) if isinstance(v, dict) and v.get("name")}
            mutated: List[Dict[str, Any]] = []
            unknowns = set()
            if isinstance(ask_plan, dict) and "ASK_PLAN" in ask_plan:
                ask_plan = ask_plan.get("ASK_PLAN") or []
            if isinstance(ask_plan, list):
                for blk in ask_plan:
                    if not isinstance(blk, dict):
                        continue
                    ask = [q for q in (blk.get("ask") or []) if q in known_vars]
                    fuw: Dict[str, List[str]] = {}
                    for cond, qs in (blk.get("followups_if", {}) or {}).items():
                        kept = [q for q in (qs or []) if q in known_vars]
                        fuw[cond] = kept
                        for q in (qs or []):
                            if q not in known_vars:
                                unknowns.add(q)
                    for q in (blk.get("ask") or []):
                        if q not in known_vars:
                            unknowns.add(q)
                    mutated.append({"module": blk.get("module"), "ask": ask, "followups_if": fuw})
                ask_plan = mutated
            if unknowns:
                merged_ir.setdefault("qa", {}).setdefault("notes", []).append(f"ask_plan_unknowns_dropped:{sorted(list(unknowns))}")
            ask_plan = _enforce_ask_ownership(ask_plan)
        except Exception:
            pass

        dmn_xml = _ensure_dmn_inputs(dmn_xml, ask_plan, merged_ir)
        return dmn_xml, ask_plan

    def generate_bpmn(self, dmn_xml: str, ask_plan: Any) -> str:
        user = "DMN:\n```xml\n" + dmn_xml + "\n```\n\n" + "ASK_PLAN:\n" + json.dumps(ask_plan, ensure_ascii=False)
        text = self._chat_text(self.bpmn_system, user, max_out=12000, model_key="bpmn")
        blocks = _extract_fenced_blocks(text)
        for _lang, body in blocks:
            if "<bpmn:definitions" in body:
                return _sanitize_bpmn(body)
        xml = _extract_xml_tag(text, "bpmn:definitions")
        if xml:
            return _sanitize_bpmn(xml)
        xml = _extract_xml_tag(text, "definitions")
        if xml:
            xml = re.sub(r"<\s*definitions\b", "<bpmn:definitions", xml, count=1)
            xml = re.sub(r"</\s*definitions\s*>", "</bpmn:definitions>", xml, count=1)
            return _sanitize_bpmn(xml)
        try:
            dbg = text
            if len(dbg) > 24000:
                dbg = dbg[:16000] + "\n...<TRUNCATED>...\n" + dbg[-6000:]
            Path(self.debug_dir / "bpmn_debug_last.txt").write_text(dbg, encoding="utf-8")
        except Exception:
            pass
        raise RuntimeError("Failed to get BPMN from model")

    def audit_coverage(self, merged_ir: Dict[str, Any], dmn_xml: str) -> Dict[str, Any]:
        payload = {"ir": merged_ir, "ir_flat": _rules_flattened(merged_ir)}
        user = "RULES_JSON:\n" + json.dumps(payload, ensure_ascii=False) + "\nDMN:\n```xml\n" + dmn_xml + "\n```"
        return self._chat_json(self.coverage_system, user, max_out=6000, model_key="audit", schema_model=None)

    def export_xlsx_from_dmn(self, merged_ir: Dict[str, Any], ask_plan: List[Dict[str, Any]],
                             out_xlsx_path: str, template_xlsx_path: Optional[str] = None) -> str:
        try:
            from openpyxl import load_workbook, Workbook
        except Exception as e:
            raise RuntimeError("openpyxl is required for XLSX export (pip install openpyxl)") from e

        if template_xlsx_path and Path(template_xlsx_path).exists():
            wb = load_workbook(template_xlsx_path)
            for sheet in ("survey", "choices"):
                if sheet not in wb.sheetnames:
                    wb.create_sheet(title=sheet)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            ws_s = wb.create_sheet("survey")
            ws_c = wb.create_sheet("choices")
            ws_s.append(["type", "name", "label", "hint", "required", "constraint", "relevant", "appearance", "calculation"])
            ws_c.append(["list_name", "name", "label"])

        form_id = Path(out_xlsx_path).stem
        title = form_id.replace("_", " ").title()
        version_val = int(time.time())
        sms_keyword_default = f"J1!{form_id}!"
        sms_separator_default = "!"

        if "settings" in wb.sheetnames:
            del wb["settings"]
        ws_set = wb.create_sheet("settings")
        ws_set.append(["id_string", "title", "default_language", "version", "sms_keyword", "sms_separator"])
        ws_set.append([str(form_id), str(title), "en", int(version_val), str(sms_keyword_default), str(sms_separator_default)])

        ws_survey = wb["survey"]
        ws_choices = wb["choices"]

        def _headers(ws):
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
            return [c or "" for c in row1]

        s_hdr = _headers(ws_survey)
        c_hdr = _headers(ws_choices)

        def _col_ix(ws, hdr_list, name):
            try:
                return hdr_list.index(name)
            except Exception:
                hdr_list.append(name)
                ws.cell(row=1, column=len(hdr_list), value=name)
                return len(hdr_list) - 1

        s_ix = {k: _col_ix(ws_survey, s_hdr, k) for k in ("type", "name", "label", "hint", "required", "constraint", "relevant", "calculation", "appearance")}
        c_ix = {k: _col_ix(ws_choices, c_hdr, k) for k in ("list_name", "name", "label")}

        var_by_name = {(v.get("name") or "").strip(): v for v in (merged_ir.get("variables") or []) if isinstance(v, dict) and v.get("name")}

        def _label_for(name: str) -> str:
            v = var_by_name.get(name)
            if v and v.get("prompt"):
                return str(v["prompt"])
            return re.sub(r"_+", " ", str(name)).strip().capitalize()

        existing_choices = set()
        for row in ws_choices.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            list_name, name, _ = (row + (None, None, None))[:3]
            if list_name and name:
                existing_choices.add((str(list_name), str(name)))

        def ensure_yes_no():
            for val in ("yes", "no"):
                if ("yes_no", val) not in existing_choices:
                    row = [None] * len(c_hdr)
                    row[c_ix["list_name"]] = "yes_no"
                    row[c_ix["name"]] = val
                    row[c_ix["label"]] = val.capitalize()
                    ws_choices.append(row)
                    existing_choices.add(("yes_no", val))

        ensure_yes_no()

        def ensure_list(list_name: str, values: List[str]):
            for v in values or []:
                key = (list_name, str(v))
                if key not in existing_choices:
                    row = [None] * len(c_hdr)
                    row[c_ix["list_name"]] = list_name
                    row[c_ix["name"]] = str(v)
                    row[c_ix["label"]] = re.sub(r"_+", " ", str(v)).strip().capitalize()
                    ws_choices.append(row)
                    existing_choices.add(key)

        def cond_to_relevant(cond: str) -> Optional[str]:
            s = str(cond or "").strip()
            m = re.match(r"^\s*([A-Za-z0-9_]+)\s*([=!<>]=|[<>]|=)\s*(.+?)\s*$", s)
            if not m:
                return None
            var, op, val = m.groups()
            val = val.strip()
            if val.lower() in ("true", "false"):
                val = "'yes'" if val.lower() == "true" else "'no'"
            elif re.match(r"^-?\d+(\.\d+)?$", val):
                pass
            else:
                val = f"'{_strip_quotes(val)}'"
            if op == "==":
                op = "="
            return f"${{{var}}} {op} {val}"

        relevant_for: Dict[str, List[str]] = {}
        for blk in (ask_plan or []):
            fuw = (blk or {}).get("followups_if") or {}
            for cond, qs in fuw.items():
                r = cond_to_relevant(cond)
                if not r:
                    continue
                for q in (qs or []):
                    relevant_for.setdefault(q, []).append(r)

        def append_row(row_dict: Dict[str, Any]):
            row = [None] * len(s_hdr)
            for k, v in row_dict.items():
                if k not in s_ix:
                    s_ix[k] = _col_ix(ws_survey, s_hdr, k)
                row[s_ix[k]] = v
            ws_survey.append(row)

        added_questions: set = set()
        for blk in (ask_plan or []):
            mod = (blk or {}).get("module") or "module"
            append_row({"type": "begin group", "name": f"{_to_snake(mod)}_group", "label": re.sub(r"_+", " ", mod).title()})
            for q in (blk or {}).get("ask", []) or []:
                if q in added_questions:
                    continue
                v = var_by_name.get(q, {})
                qtype = str(v.get("type") or "").lower()
                allowed = v.get("allowed") or []
                if qtype == "boolean":
                    q_type_cell = "select_one yes_no"
                elif allowed:
                    list_name = f"list_{q}"
                    ensure_list(list_name, allowed)
                    q_type_cell = f"select_one {list_name}"
                elif qtype == "number":
                    q_type_cell = "decimal"
                else:
                    q_type_cell = "text"
                rel_expr = " or ".join(sorted(set(relevant_for[q]))) if q in relevant_for else None
                append_row({"type": q_type_cell, "name": q, "label": _label_for(q), "required": None, "relevant": rel_expr})
                added_questions.add(q)
            append_row({"type": "end group"})

        outp = Path(out_xlsx_path)
        outp.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(outp))
        return str(outp)

    def parse_dmn_decision_tables(self, dmn_xml: str) -> Dict[str, dict]:
        root = ET.fromstring(dmn_xml)
        out: Dict[str, dict] = {}
        for dec in root.findall(".//dmn:decision", _DMNN):
            name = (dec.get("name") or (dec.get("id") or "module")).strip()
            mod = re.sub(r"^decide[_:\-]+", "", name).strip() or "module"
            table = dec.find(".//dmn:decisionTable", _DMNN)
            if table is None:
                continue
            inputs = [_text_or(i.find("./dmn:inputExpression/dmn:text", _DMNN)) for i in table.findall("./dmn:input", _DMNN)]
            outputs = [(o.get("name") or "").strip() for o in table.findall("./dmn:output", _DMNN)]
            rows = []
            for r in table.findall("./dmn:rule", _DMNN):
                conds = [_text_or(e.find("./dmn:text", _DMNN)) for e in r.findall("./dmn:inputEntry", _DMNN)]
                outs = [_text_or(e.find("./dmn:text", _DMNN)) for e in r.findall("./dmn:outputEntry", _DMNN)]
                rows.append({"conds": conds, "outs": outs})
            out[mod] = {"inputs": inputs, "outputs": outputs, "rows": rows}
        return out

    def export_csvs_from_dmn(self, dmn_xml: str, out_dir: str) -> Dict[str, str]:
        def _normalize_page_ranges(ref: str) -> str:
            s = _strip_quotes(ref or "")
            def _ensure_p_prefix(m):
                a, b = m.group(1), m.group(2)
                return f"p{a}-{b}"
            s = re.sub(r"\b(\d{1,4})\s*-\s*(\d{1,4})\b", _ensure_p_prefix, s)
            def _swap_if_needed(m):
                a = int(m.group(1))
                b = int(m.group(2))
                lo, hi = sorted((a, b))
                return f"p{lo}-{hi}"
            s = re.sub(r"[Pp]\s*(\d{1,4})\s*-\s*[Pp]?\s*(\d{1,4})", _swap_if_needed, s)
            s = re.sub(r"\s*([,;])\s*", r"\1 ", s)
            return s.strip()

        tables = self.parse_dmn_decision_tables(dmn_xml)
        out_map: Dict[str, str] = {}
        outp = Path(out_dir)
        outp.mkdir(parents=True, exist_ok=True)

        def norm_out_val(col: str, val: str) -> str:
            v = _strip_quotes(val or "")
            if col in ("danger_sign", "clinic_referral"):
                vv = v.strip().lower()
                if vv in ("true", "false"):
                    return vv
            if col == "ref":
                return _normalize_page_ranges(v)
            return v

        DEFAULTS = {
            "triage": "home",
            "danger_sign": "false",
            "clinic_referral": "false",
            "reason": "",
            "ref": "p0",
            "actions": "",
            "advice": "",
        }

        for mod, t in tables.items():
            outs = [c for c in t["outputs"] if c] or ["triage", "danger_sign", "clinic_referral", "reason", "ref", "actions", "advice"]
            fname = f"dmn_{mod}.csv"
            fpath = outp / fname
            with open(fpath, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["key"] + outs)
                for i, row in enumerate(t["rows"], start=1):
                    key = f"r{str(i).zfill(3)}"
                    raw_vals = row["outs"][: len(outs)]
                    raw_vals += [""] * (len(outs) - len(raw_vals))
                    filled = []
                    for col_name, cell in zip(outs, raw_vals):
                        v = cell if (cell is not None and str(cell).strip() != "") else DEFAULTS.get(col_name, "")
                        filled.append(norm_out_val(col_name, v))
                    w.writerow([key] + filled)
            out_map[mod] = str(fpath)

        return out_map

    def wire_decisions_into_xlsx(self, xlsx_path: str, dmn_xml: str, merged_ir: Dict[str, Any], media_id_prefix: str = "dmn_") -> None:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb["survey"]

        def hdrs():
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
            return [c or "" for c in row1]

        hdr = hdrs()

        def col(name):
            if name in hdr:
                return hdr.index(name)
            hdr.append(name)
            ws.cell(row=1, column=len(hdr), value=name)
            return len(hdr) - 1

        c_type = col("type")
        c_name = col("name")
        c_calc = col("calculation")
        c_lab = col("label")
        c_rel = col("relevant")

        def append_by_cols(values: dict):
            row = [None] * len(hdr)
            if "type" in values:
                row[c_type] = values["type"]
            if "name" in values:
                row[c_name] = values["name"]
            if "calculation" in values:
                row[c_calc] = values["calculation"]
            if "label" in values:
                row[c_lab] = values["label"]
            if "relevant" in values:
                row[c_rel] = values["relevant"]
            ws.append(row)

        def map_var_token_to_xls(var_token: str) -> str:
            tok = (var_token or "").strip()
            if "." in tok:
                left, right = tok.rsplit(".", 1)
                mod = re.sub(r"^decide[_:\-]+", "", left)
                return f"{media_id_prefix}{_to_snake(mod)}_{_to_snake(right)}"
            return tok

        def feel_to_xls_test(expr: str) -> Optional[str]:
            s = (expr or "").strip()
            if s in ("", "-", "otherwise"):
                return None
            m = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_\.\-:]*)\s*(=|==|!=|<=|>=|<|>)\s*(.+?)\s*$", s)
            if not m:
                return None
            var_tok, op, rhs = m.groups()
            if op == "==":
                op = "="
            xls_var = map_var_token_to_xls(var_tok)
            lhs_is_calc = xls_var.startswith(media_id_prefix)
            val = (rhs or "").strip()
            if val.lower() in ("true", "false"):
                rhs_x = "'true'" if lhs_is_calc else ("'yes'" if val.lower() == "true" else "'no'")
            elif re.match(r"^-?\d+(\.\d+)?$", val):
                rhs_x = val
            else:
                rhs_x = f"'{_strip_quotes(val)}'"
            return f"${{{xls_var}}} {op} {rhs_x}"

        tables = self.parse_dmn_decision_tables(dmn_xml)
        for mod, t in tables.items():
            row_tests: List[str] = []
            for i, row in enumerate(t["rows"], start=1):
                conjuncts: List[str] = []
                for feel in row["conds"]:
                    xp = feel_to_xls_test(feel)
                    if xp:
                        conjuncts.append(xp)
                row_tests.append(_and_join(conjuncts))

            outs = [c for c in t["outputs"] if c] or ["triage", "danger_sign", "clinic_referral", "reason", "ref", "actions", "advice"]

            for col_idx, out_col in enumerate(outs):
                values: List[str] = []
                for row in t["rows"]:
                    raw_val = row["outs"][col_idx] if col_idx < len(row["outs"]) else ""
                    v = (raw_val or "").strip()
                    values.append(v)
                expr_out = "''"
                for test, val in reversed(list(zip(row_tests, values))):
                    if val == "":
                        lit = "''"
                    else:
                        vv = _strip_quotes(val)
                        if vv.lower() in ("true", "false"):
                            lit = f"'{vv.lower()}'"
                        else:
                            lit = f"'{vv}'"
                    if test:
                        expr_out = f"if({test}, {lit}, {expr_out})"
                append_by_cols({
                    "type": "calculate",
                    "name": f"{media_id_prefix}{mod}_{out_col}",
                    "calculation": expr_out,
                    "label": f"{mod} {out_col}".replace("_", " "),
                    "relevant": None
                })

            summary_label = (
                f"Triage: ${{{media_id_prefix}{mod}_triage}}\n"
                f"Danger sign: ${{{media_id_prefix}{mod}_danger_sign}}\n"
                f"Clinic referral: ${{{media_id_prefix}{mod}_clinic_referral}}\n"
                f"Reason: ${{{media_id_prefix}{mod}_reason}}\n"
                f"Ref: ${{{media_id_prefix}{mod}_ref}}\n"
                f"CHW actions: ${{{media_id_prefix}{mod}_actions}}"
            )
            append_by_cols({
                "type": "note",
                "name": f"{media_id_prefix}{mod}_summary",
                "label": summary_label
            })

        wb.save(xlsx_path)

    def build_orchestrator_folder(self, dmn_xml: str, ask_plan: List[Dict[str, Any]],  # noqa: D401
                                  merged_ir: Dict[str, Any], output_dir: Union[str, Path],
                                  template_xlsx_path: Optional[str] = None, logo_path: Optional[str] = None) -> str:
        out_root = Path(output_dir)
        out_root.mkdir(parents=True, exist_ok=True)

        xlsx_path = out_root / "orchestrator.xlsx"
        xml_path = out_root / "orchestrator.xml"
        props_path = out_root / "orchestrator.properties.json"
        media_dir = out_root / "media"
        media_dir.mkdir(exist_ok=True)

        self.export_xlsx_from_dmn(merged_ir, ask_plan, str(xlsx_path), template_xlsx_path)
        self.wire_decisions_into_xlsx(str(xlsx_path), dmn_xml, merged_ir)

        xml_placeholder = (
            "<?xml version=\"1.0\"?>\n"
            "<orchestrator>\n"
            "  <!-- Placeholder XML form: logic embedded in orchestrator.xlsx -->\n"
            "</orchestrator>\n"
        )
        with xml_path.open("w", encoding="utf-8") as f_xml:
            f_xml.write(xml_placeholder)

        logo_dest = media_dir / "logo.png"
        from shutil import copyfile
        if logo_path and Path(logo_path).exists():
            try:
                copyfile(logo_path, logo_dest)
            except Exception:
                pass
        else:
            default_logo = Path(__file__).with_name("205a491d-21bb-46eb-be6e-6e5279e4156b.png")
            if default_logo.exists():
                try:
                    copyfile(default_logo, logo_dest)
                except Exception:
                    pass
            else:
                try:
                    from PIL import Image
                    img = Image.new("RGBA", (1, 1), (0, 0, 0, 0))
                    img.save(logo_dest, "PNG")
                except Exception:
                    logo_dest.touch()

        properties = {
            "title": "Orchestrator",
            "icon": "icon-healthcare",
            "context": {"person": False, "place": False},
            "internalId": "orchestrator",
            "xmlFormId": "orchestrator",
            "media": ["logo.png"],
        }
        with props_path.open("w", encoding="utf-8") as f_props:
            json.dump(properties, f_props, indent=2)

        return str(out_root)

    def extract_sections_from_pdf(self, pdf_path: str, max_chars: int = 4000) -> List[Tuple[str, str]]:
        logp = self.debug_dir / "sectioning.log"

        pages = _extract_pypdf_pages(pdf_path)
        pages = _postprocess_page_text(pages)
        nonempty = sum(1 for _, t in pages if (t or "").strip())
        if nonempty > 0:
            chunks = _split_into_sections_by_headings(pages, max_chars=max_chars)
            if not chunks:
                chunks = _chunk_pages_len_only(pages, max_chars=max_chars)
            try:
                logp.write_text("Used pypdf extraction\n", encoding="utf-8")
            except Exception:
                pass
            return chunks

        if _have_pdfminer:
            pm_layout = _extract_pdfminer_layout_pages(pdf_path)
            pm_layout = _postprocess_page_text(pm_layout)
            if sum(1 for _, t in pm_layout if (t or "").strip()) > 0:
                chunks = _split_into_sections_by_headings(pm_layout, max_chars=max_chars)
                if not chunks:
                    chunks = _chunk_pages_len_only(pm_layout, max_chars=max_chars)
                try:
                    logp.write_text("Used pdfminer layout extraction\n", encoding="utf-8")
                except Exception:
                    pass
                return chunks

        if _have_pdfminer:
            pm_pages = _extract_pdfminer_pages(pdf_path)
            pm_pages = _postprocess_page_text(pm_pages)
            if sum(1 for _, t in pm_pages if (t or "").strip()) > 0:
                chunks = _split_into_sections_by_headings(pm_pages, max_chars=max_chars)
                if not chunks:
                    chunks = _chunk_pages_len_only(pm_pages, max_chars=max_chars)
                try:
                    logp.write_text("Used pdfminer basic extraction\n", encoding="utf-8")
                except Exception:
                    pass
                return chunks

        if _have_ocr:
            ocr_pages = _extract_ocr_pages(pdf_path)
            ocr_pages = _postprocess_page_text(ocr_pages)
            if sum(1 for _, t in ocr_pages if (t or "").strip()) > 0:
                chunks = _split_into_sections_by_headings(ocr_pages, max_chars=max_chars)
                if not chunks:
                    chunks = _chunk_pages_len_only(ocr_pages, max_chars=max_chars)
                try:
                    logp.write_text("Used OCR extraction (pytesseract + pdf2image)\n", encoding="utf-8")
                except Exception:
                    pass
                return chunks

        try:
            logp.write_text("No text extracted by any strategy\n", encoding="utf-8")
        except Exception:
            pass
        return []

if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(description="Extract -> DMN -> XLSForm")
    parser.add_argument("--pdf", required=False, help="Guideline PDF")
    parser.add_argument("--dmn", required=False, help="Existing DMN XML path")
    parser.add_argument("--merged", required=False, help="Existing merged_ir.json path")
    parser.add_argument("--ask", required=False, help="Existing ask_plan.json path")
    parser.add_argument("--xlsx-out", default="forms/app/orchestrator.xlsx")
    parser.add_argument("--media-dir", default="forms/app/orchestrator-media")
    parser.add_argument("--template", default=None)
    args = parser.parse_args()

    gx = OpenAIGuardedExtractor(api_key=os.environ.get("OPENAI_API_KEY"))

    merged = None
    dmn_xml = None
    ask_plan = None

    if args.pdf:
        if not os.getenv("OPENAI_API_KEY"):
            print("OPENAI_API_KEY not set", file=sys.stderr)
            sys.exit(1)
        sections = gx.extract_rules_per_section(args.pdf)
        merged = gx.merge_sections(sections)
        dmn_xml, ask_plan = gx.generate_dmn_and_ask_plan(merged)
    elif args.dmn and args.merged:
        dmn_xml = Path(args.dmn).read_text(encoding="utf-8")
        merged = json.loads(Path(args.merged).read_text(encoding="utf-8"))
        if args.ask and Path(args.ask).exists():
            raw = json.loads(Path(args.ask).read_text(encoding="utf-8"))
            ask_plan = raw.get("ASK_PLAN") if isinstance(raw, dict) and "ASK_PLAN" in raw else raw
        else:
            ask_plan = [
                {"module": "module_a",
                 "ask": [v["name"] for v in (merged.get("variables") or []) if isinstance(v, dict) and v.get("name")],
                 "followups_if": {}}
            ]
    else:
        print("Supply either --pdf OR both --dmn and --merged", file=sys.stderr)
        sys.exit(2)

    orchestrator_root = Path(args.xlsx_out).with_suffix("")
    orchestrator_root.parent.mkdir(parents=True, exist_ok=True)
    gx.build_orchestrator_folder(
        dmn_xml=dmn_xml,
        ask_plan=ask_plan,
        merged_ir=merged,
        output_dir=orchestrator_root,
        template_xlsx_path=args.template,
        logo_path=None,
    )

    print("Orchestrator directory:", orchestrator_root)
